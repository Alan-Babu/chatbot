from fastapi import FastAPI, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sentence_transformers import SentenceTransformer
from pydantic import BaseModel
import faiss, os, re, logging, hashlib
from dotenv import load_dotenv
from typing import Generator, Optional

# File readers
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
from sklearn.preprocessing import normalize
from sklearn.feature_extraction.text import TfidfVectorizer

import ollama
from datetime import datetime
from rapidfuzz import process as fuzz_process

# -------------------------
# Config
# -------------------------
load_dotenv()
app = FastAPI()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EMBEDDER_MODEL = os.getenv("EMBEDDER_MODEL", "all-MiniLM-L6-v2")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "phi3")

embedder = SentenceTransformer(EMBEDDER_MODEL)

# Use app.state instead of globals
app.state.index = None
app.state.docs = []
app.state.doc_sources = []

class ChatRequest(BaseModel):
    query: str
    k: int = 3
    session_id: Optional[str] = None


# -------------------------
# File loaders
# -------------------------
def load_txt(file_path): 
    with open(file_path, "r", encoding="utf-8") as f: 
        return f.read()

def load_md(file_path): 
    return load_txt(file_path)

def load_pdf(file_path):
    reader = PdfReader(file_path)
    texts = [page.extract_text() for page in reader.pages if page.extract_text()]
    return "\n".join(texts)

def load_docx(file_path):
    doc = Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def load_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text


# -------------------------
# Chunking
# -------------------------
def chunk_text(text, chunk_size=500, overlap=50):
    """Split text into chunks with overlap."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks


# -------------------------
# Startup: Scan folder + index
# -------------------------
def build_index_from_dir(file_dir: str):
    """Load files from a directory and build/update FAISS index."""
    app.state.docs = []
    app.state.doc_sources = []

    if not os.path.exists(file_dir):
        logger.warning("⚠️ Data folder not found.")
        return

    for file in os.listdir(file_dir):
        file_path = os.path.join(file_dir, file)
        ext = os.path.splitext(file)[1].lower()

        try:
            if ext == ".txt": text = load_txt(file_path)
            elif ext == ".md": text = load_md(file_path)
            elif ext == ".pdf": text = load_pdf(file_path)
            elif ext == ".docx": text = load_docx(file_path)
            elif ext == ".xlsx": text = load_xlsx(file_path)
            else:
                logger.info(f"Skipping unsupported file: {file}")
                continue

            chunks = chunk_text(text)
            app.state.docs.extend(chunks)
            app.state.doc_sources.extend([file] * len(chunks))
            logger.info(f"✅ Loaded {file} with {len(chunks)} chunks")

        except Exception as e:
            logger.error(f"❌ Failed to load {file}: {e}")

    if not app.state.docs:
        logger.warning("⚠️ No documents loaded.")
        return

    embeddings = embedder.encode(app.state.docs)
    embeddings = normalize(embeddings)
    dim = embeddings.shape[1]
    app.state.index = faiss.IndexFlatIP(dim)
    app.state.index.add(embeddings)

    logger.info(
        f"📚 Index built with {len(app.state.docs)} chunks from {len(set(app.state.doc_sources))} files"
    )


@app.on_event("startup")
def load_data():
    file_dir = os.getenv("DATA_DIR", "data/")
    build_index_from_dir(file_dir)
    # Initialize DB
    try:
        from .db import init_db
        init_db()
    except Exception as e:
        logger.warning(f"DB init failed: {e}")


# -------------------------
# Chat Endpoint (RAG with ollama)
# -------------------------
@app.post("/chat")
def chat(request: ChatRequest):
    if app.state.index is None:
        raise HTTPException(status_code=500, detail="Index not initialized")

    # Basic cache via hash key
    cache_key = hashlib.sha256(f"{request.query.strip().lower()}::{request.k}".encode()).hexdigest()
    cached_response = None
    try:
        from .db import SessionLocal, CacheEntry, ChatMessage
        db = SessionLocal()
        try:
            cached_entry = db.query(CacheEntry).filter(CacheEntry.query_key == cache_key).first()
            if cached_entry and not cached_entry.is_expired:
                cached_response = cached_entry.response
                if request.session_id:
                    db.add(ChatMessage(session_id=request.session_id, role='user', content=request.query))
                    db.add(ChatMessage(session_id=request.session_id, role='assistant', content=cached_response))
                    db.commit()
        finally:
            db.close()
    except Exception:
        pass

    if cached_response is not None:
        def gen_cached():
            yield cached_response
        return StreamingResponse(gen_cached(), media_type="text/plain")

    q_emb = embedder.encode([request.query])
    q_emb = normalize(q_emb)  # normalize query too

    D, I = app.state.index.search(q_emb, request.k)
    retrieved_chunks = [(app.state.docs[i], app.state.doc_sources[i]) for i in I[0]]

    context = "\n".join([f"From {src}:\n{chunk}" for chunk, src in retrieved_chunks])

    prompt = f"""You are a helpful assistant.
Use the following context to answer the question. do not tell any information about the datasets or documents used to answer the question.

Context:
{context}

Question: {request.query}
Answer:"""

    def generate() -> Generator[str, None, None]:
        try:
            for chunk in ollama.chat(
                model=OLLAMA_MODEL,
                messages=[
                    {"role": "system", "content": "You are a knowledgeable assistant."},
                    {"role": "user", "content": prompt}
                ],
                stream=True
            ):
                if chunk.get("message") and "content" in chunk["message"]:
                    yield chunk["message"]["content"]
        except Exception as e:
            yield f"Error generating response: {str(e)}"

    def aggregate_and_stream():
        full = []
        for piece in generate():
            full.append(piece)
            yield piece
        try:
            from .db import SessionLocal, CacheEntry, ChatMessage
            db = SessionLocal()
            try:
                text = "".join(full)
                entry = db.query(CacheEntry).filter(CacheEntry.query_key == cache_key).first()
                if entry is None:
                    entry = CacheEntry(query_key=cache_key, response=text, ttl_seconds=int(os.getenv("CACHE_TTL", "3600")))
                    db.add(entry)
                else:
                    entry.response = text
                    entry.created_at = datetime.utcnow()
                if request.session_id:
                    db.add(ChatMessage(session_id=request.session_id, role='user', content=request.query))
                    db.add(ChatMessage(session_id=request.session_id, role='assistant', content=text))
                db.commit()
            finally:
                db.close()
        except Exception:
            pass

    return StreamingResponse(aggregate_and_stream(), media_type="text/plain")


# -------------------------
# Menu Endpoint
# -------------------------
@app.get("/menu", response_model=list[str])
def get_menu():
    if not app.state.doc_sources:
        raise HTTPException(status_code=404, detail="No documents loaded")

    menu_items = []
    for doc in app.state.docs:
        headings = re.findall(r"^(#+\s.*)", doc, flags=re.MULTILINE)
        if headings:
            menu_items.extend([h.strip("# ").strip() for h in headings])

    if not menu_items:
        vectorizer = TfidfVectorizer(stop_words="english", max_features=10)
        X = vectorizer.fit_transform(app.state.docs)
        keywords = vectorizer.get_feature_names_out()
        menu_items.extend(keywords.tolist())

    # Deduplicate and return top 15
    unique_items = list(dict.fromkeys(menu_items))[:15]
    return unique_items


# -------------------------
# Feedback, History, Fuzzy Search, Suggestions
# -------------------------
@app.post("/feedback/message")
def feedback_message(payload: dict):
    try:
        from .db import SessionLocal, MessageFeedback
        db = SessionLocal()
        try:
            message_id = payload.get("messageId")
            feedback = payload.get("feedback")
            if not message_id or feedback not in ("up", "down"):
                raise HTTPException(status_code=400, detail="Invalid payload")
            db.add(MessageFeedback(message_id=message_id, feedback=feedback))
            db.commit()
            return {"success": True}
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/feedback/session")
def feedback_session(payload: dict):
    try:
        from .db import SessionLocal, SessionFeedback
        db = SessionLocal()
        try:
            rating = payload.get("rating")
            if not isinstance(rating, int) or rating < 1 or rating > 5:
                raise HTTPException(status_code=400, detail="Invalid rating")
            db.add(SessionFeedback(rating=rating))
            db.commit()
            return {"success": True}
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/history/{session_id}")
def history(session_id: str):
    try:
        from .db import SessionLocal, ChatMessage
        db = SessionLocal()
        try:
            rows = (
                db.query(ChatMessage)
                .filter(ChatMessage.session_id == session_id)
                .order_by(ChatMessage.created_at.asc())
                .all()
            )
            return [
                {"role": r.role, "content": r.content, "timestamp": r.created_at.isoformat()}
                for r in rows
            ]
        finally:
            db.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/search")
def fuzzy_search(q: str, limit: int = 5):
    corpus = app.state.docs or []
    if not corpus:
        raise HTTPException(status_code=404, detail="No documents loaded")
    matches = fuzz_process.extract(q, corpus, limit=limit)
    # matches: list of tuples (text, score, index)
    out = []
    for text, score, idx in matches:
        out.append({
            "snippet": text[:300],
            "score": float(score),
            "source": app.state.doc_sources[idx] if idx < len(app.state.doc_sources) else None
        })
    return out


@app.post("/suggestions")
def suggestions(payload: dict):
    # naive rule-based suggestions; can be replaced with LLM or templates
    text = (payload.get("text") or "").lower()
    sugg = []
    if any(k in text for k in ["register", "signup", "sign up", "account"]):
        sugg += ["Steps to register", "Required documents", "Common issues"]
    if any(k in text for k in ["payment", "fee", "invoice", "billing"]):
        sugg += ["How to pay", "Verify payment", "Refund process"]
    if any(k in text for k in ["login", "uae pass", "access"]):
        sugg += ["Reset UAE Pass", "Link company", "Browser issues"]
    if not sugg:
        sugg = ["Ask another question", "Upload a document", "See menu topics"]
    return {"suggestions": sugg[:5]}


# -------------------------
# Healthcheck Endpoint
# -------------------------
@app.get("/health")
def health():
    return {
        "status": "ok",
        "docs_loaded": len(app.state.docs),
        "files_loaded": len(set(app.state.doc_sources)),
        "embedder_model": EMBEDDER_MODEL,
        "ollama_model": OLLAMA_MODEL,
    }


# -------------------------
# Ingest Endpoint: rebuild index from DATA_DIR
# -------------------------
@app.post("/ingest")
def ingest():
    data_dir = os.getenv("DATA_DIR", "data/")
    build_index_from_dir(data_dir)
    return {
        "status": "ok",
        "docs_loaded": len(app.state.docs),
        "files_loaded": len(set(app.state.doc_sources)),
    }
